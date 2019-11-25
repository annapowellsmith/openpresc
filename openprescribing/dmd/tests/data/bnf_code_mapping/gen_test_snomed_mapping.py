# This script takes a BNF code mapping spreadsheet, and generates a new
# spreadsheet with the subset of the rows corresponding to VMP/AMP/VMPP/AMPP
# records in the database.
#
# Usage:
#
#   python gen_test_snomed_mapping.py [inp_path] [outp_path]


import os
import sys

from openpyxl import Workbook, load_workbook
import psycopg2

if len(sys.argv) != 3:
    print("Usage: python gen_test_snomed_mapping.py [inp_path] [outp_path]")
    sys.exit(1)

inp_path = sys.argv[1]
outp_path = sys.argv[2]

connection = psycopg2.connect(
    database=os.getenv("DB_NAME"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASS"),
)

cursor = connection.cursor()

keys = set()

cursor.execute("SELECT vpid FROM dmd_vmp")
keys |= {("VMP", row[0]) for row in cursor.fetchall()}

cursor.execute("SELECT vppid FROM dmd_vmpp")
keys |= {("VMPP", row[0]) for row in cursor.fetchall()}

cursor.execute("SELECT apid FROM dmd_amp")
keys |= {("AMP", row[0]) for row in cursor.fetchall()}

cursor.execute("SELECT appid FROM dmd_ampp")
keys |= {("AMPP", row[0]) for row in cursor.fetchall()}

wb_in = load_workbook(inp_path)
wb_out = Workbook()

rows = wb_in.active.rows

headers = next(rows)
assert headers[1].value == "VMP / VMPP/ AMP / AMPP"
assert headers[2].value == "BNF Code"
assert headers[4].value == "SNOMED Code"

ws = wb_out.active
header_values = [cell.value for cell in headers]
ws.append(header_values)

for row in rows:
    snomed_code = row[4].value

    if not snomed_code:
        continue

    if snomed_code[0] == "'":
        snomed_code = snomed_code[1:]
        if not snomed_code:
            continue

    key = (row[1].value, int(snomed_code))

    if key in keys:
        row_values = [cell.value for cell in row]
        ws.append(row_values)

wb_out.save(outp_path)
